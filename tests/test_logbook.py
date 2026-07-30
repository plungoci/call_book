import tempfile
import unittest
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from unittest import mock

from call_book.adif_export import adif_record
from call_book.database import Database
from call_book.excel_export import export_excel
from call_book.models import QSO, OperatorProfile, Repeater
from call_book.services.location_service import (
    IpLocationService,
    LocationDnsError,
    LocationResponseError,
    LocationService,
)
from call_book.ui.main_window import qso_table_dates
from call_book.utils.maidenhead import coordinates_to_maidenhead, maidenhead_to_coordinates
from call_book.validators import (
    band_for_frequency,
    format_name_input,
    normalize_callsign,
    normalize_name,
    parse_positive,
    validate_callsign,
    validate_qso,
)


class LogbookTests(unittest.TestCase):
    def qso(self, **changes: Any) -> QSO:
        values: dict[str, Any] = dict(
            callsign="yo3abc/p",
            qso_start_utc="2026-01-01T12:00:00+00:00",
            qso_end_utc="2026-01-01T12:01:00+00:00",
            frequency_mhz=145.5,
            mode="FM",
        )
        values.update(changes)
        return QSO(**values)

    def test_callsign_normalization(self):
        self.assertEqual(validate_callsign(" yo3abc/m "), "YO3ABC/M")

    def test_live_callsign_formatting(self):
        self.assertEqual(normalize_callsign("yo8abc/p"), "YO8ABC/P")

    def test_name_formatting(self):
        self.assertEqual(normalize_name("  iON   popescu "), "Ion Popescu")
        self.assertEqual(format_name_input("mihai  "), "Mihai ")

    def test_invalid_frequency(self):
        with self.assertRaises(ValueError):
            parse_positive("0", "Frecvența")

    def test_bands(self):
        self.assertEqual(band_for_frequency(145.5), "2m")
        self.assertEqual(band_for_frequency(999), "Unknown")

    def test_time_interval(self):
        with self.assertRaises(ValueError):
            validate_qso(self.qso(qso_end_utc="2026-01-01T11:00:00+00:00"))

    def test_qso_table_dates_show_local_time_and_utc_values(self):
        local_time, utc_date, utc_time = qso_table_dates("2026-01-01T23:15:30+00:00")
        self.assertEqual(utc_date, "2026-01-01")
        self.assertEqual(utc_time, "23:15:30")
        self.assertEqual(local_time, datetime(2026, 1, 1, 23, 15, 30, tzinfo=UTC).astimezone().strftime("%H:%M:%S"))

    def test_adif_byte_lengths(self):
        record = adif_record(validate_qso(self.qso(notes="ș")))
        self.assertIn("<COMMENT:2>ș", record)
        self.assertIn("<EOR>", record)

    def test_crud_and_duplicate(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = Database(Path(tmp) / "test.db")
            q = validate_qso(self.qso())
            ident = db.save_qso(q)
            self.assertEqual(db.get_qso(ident).callsign, "YO3ABC/P")
            self.assertTrue(db.possible_duplicate(q))
            q.id = ident
            q.notes = "edited"
            db.save_qso(q)
            self.assertEqual(db.get_qso(ident).notes, "edited")
            r = Repeater("R0", 145.6)
            rid = db.save_repeater(r)
            q.repeater_id = rid
            db.save_qso(q)
            db.delete_repeater(rid)
            self.assertIsNone(db.get_qso(ident).repeater_id)
            db.delete_qso(ident)
            self.assertEqual(db.list_qsos(), [])

    def test_reset_id_sequences_starts_empty_tables_at_one(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = Database(Path(tmp) / "test.db")
            q = validate_qso(self.qso())
            ident = db.save_qso(q)
            db.delete_qso(ident)
            db.reset_id_sequences()
            self.assertEqual(db.save_qso(q), 1)

    def test_reset_id_sequences_preserves_existing_records(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = Database(Path(tmp) / "test.db")
            first = db.save_qso(validate_qso(self.qso()))
            second = db.save_qso(
                validate_qso(
                    self.qso(qso_start_utc="2026-01-02T12:00:00+00:00", qso_end_utc="2026-01-02T12:01:00+00:00")
                )
            )
            db.delete_qso(second)
            db.reset_id_sequences()
            self.assertEqual(
                db.save_qso(
                    validate_qso(
                        self.qso(qso_start_utc="2026-01-03T12:00:00+00:00", qso_end_utc="2026-01-03T12:01:00+00:00")
                    )
                ),
                first + 1,
            )

    def test_propagation_crud_and_delete(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = Database(Path(tmp) / "test.db")
            q = validate_qso(
                self.qso(
                    propagation_mode="Satelit",
                    satellite_name="QO-100",
                    uplink_mode="SSB",
                    downlink_mode="SSB",
                    distance_km=35786.5,
                    azimuth_deg=180,
                    propagation_notes="QSB",
                )
            )
            ident = db.save_qso(q)
            saved = db.get_qso(ident)
            self.assertEqual((saved.satellite_name, saved.distance_km, saved.azimuth_deg), ("QO-100", 35786.5, 180))
            saved.propagation_notes = "edited"
            db.save_qso(saved)
            self.assertEqual(db.get_qso(ident).propagation_notes, "edited")
            db.delete_qso(ident)
            with self.assertRaises(KeyError):
                db.get_qso(ident)

    def test_propagation_validation(self):
        with self.assertRaises(ValueError):
            validate_qso(self.qso(distance_km=0))
        with self.assertRaises(ValueError):
            validate_qso(self.qso(azimuth_deg=361))
        with self.assertRaises(ValueError):
            validate_qso(self.qso(propagation_mode="Satelit"))
        self.assertEqual(
            validate_qso(
                self.qso(propagation_mode="Satelit", satellite_name="QO-100", uplink_mode="SSB", downlink_mode="SSB")
            ).satellite_name,
            "QO-100",
        )

    def test_propagation_exports(self):
        q = validate_qso(
            self.qso(
                propagation_mode="Satelit",
                satellite_name="QO-100",
                uplink_mode="SSB",
                downlink_mode="USB",
                distance_km=35786.5,
                azimuth_deg=145,
                propagation_notes="Deschidere excelentă",
            )
        )
        record = adif_record(q)
        self.assertIn("<PROP_MODE:3>SAT", record)
        self.assertIn("<SAT_NAME:6>QO-100", record)
        self.assertIn("<SAT_MODE:7>SSB/USB", record)
        self.assertIn("<DISTANCE:7>35786.5", record)
        self.assertIn("Deschidere excelentă", record)
        with tempfile.TemporaryDirectory() as tmp:
            from openpyxl import load_workbook

            path = export_excel([q], destination=Path(tmp) / "qsos.xlsx")
            ws = load_workbook(path).active
            self.assertEqual(ws[1][13].value, "Propagare")
            self.assertEqual(ws[2][13].value, "Satelit")
            self.assertEqual(ws[2][18].value, 145)

    def test_operator_profile_save_and_load(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = Database(Path(tmp) / "test.db")
            profile = OperatorProfile(
                callsign="YO3ABC", full_name="Ion Popescu", default_power_w=25.0, radio_club="Radio Club"
            )
            db.save_operator_profile(profile)
            loaded = db.get_operator_profile()
            self.assertEqual(loaded.callsign, "YO3ABC")
            self.assertEqual(loaded.full_name, "Ion Popescu")
            self.assertEqual(loaded.default_power_w, 25.0)

    def test_profile_location_save_and_load(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = Database(Path(tmp) / "test.db")
            db.save_operator_profile(
                OperatorProfile(
                    latitude=44.4268,
                    longitude=26.1025,
                    location_accuracy_m=20,
                    location_source="Windows Location",
                    location_updated_at="2026-01-01T00:00:00+00:00",
                    grid_square="KN34BK",
                )
            )
            p = db.get_operator_profile()
            self.assertEqual((p.latitude, p.longitude, p.grid_square), (44.4268, 26.1025, "KN34BK"))

    def test_existing_database_migration(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "old.db"
            import sqlite3

            c = sqlite3.connect(path)
            c.execute("CREATE TABLE operator_profile (id INTEGER PRIMARY KEY, callsign TEXT)")
            c.execute("INSERT INTO operator_profile VALUES (1,'YO3OLD')")
            c.commit()
            c.close()
            db = Database(path)
            self.assertEqual(db.get_operator_profile().callsign, "YO3OLD")
            with db.connect() as c:
                self.assertIn("latitude", [r["name"] for r in c.execute("PRAGMA table_info(operator_profile)")])

    def test_qso_propagation_migration(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "old.db"
            import sqlite3

            c = sqlite3.connect(path)
            c.execute(
                "CREATE TABLE qsos (id INTEGER PRIMARY KEY, callsign TEXT NOT NULL, "
                "qso_start_utc TEXT NOT NULL, frequency_mhz REAL NOT NULL, mode TEXT NOT NULL)"
            )
            c.commit()
            c.close()
            db = Database(path)
            with db.connect() as c:
                columns = {r["name"] for r in c.execute("PRAGMA table_info(qsos)")}
            self.assertTrue(
                {
                    "propagation_mode",
                    "satellite_name",
                    "uplink_mode",
                    "downlink_mode",
                    "distance_km",
                    "azimuth_deg",
                    "propagation_notes",
                }
                <= columns
            )

    def test_maidenhead_precisions_and_hemispheres(self):
        self.assertEqual(coordinates_to_maidenhead(44.4268, 26.1025, 4), "KN34")
        self.assertEqual(coordinates_to_maidenhead(44.4268, 26.1025, 6), "KN34bk")
        self.assertEqual(coordinates_to_maidenhead(44.4268, 26.1025, 8), "KN34bk22")
        self.assertEqual(coordinates_to_maidenhead(0, 0), "JJ00aa")
        self.assertEqual(coordinates_to_maidenhead(-33.8688, -151.2093), "BF46jd")

    def test_maidenhead_limits_invalid_and_inverse(self):
        for lat, lon in ((90, 180), (-90, -180), (90, -180), (-90, 180)):
            self.assertEqual(len(coordinates_to_maidenhead(lat, lon)), 6)
        for lat, lon in ((91, 0), (-91, 0), (0, 181), (0, -181)):
            with self.assertRaises(ValueError):
                coordinates_to_maidenhead(lat, lon)
        center_lat, center_lon = maidenhead_to_coordinates("KN34bk")
        self.assertAlmostEqual(center_lat, 44.4375)
        self.assertAlmostEqual(center_lon, 26.125)
        with self.assertRaises(ValueError):
            maidenhead_to_coordinates("ZZ99zz")

    def test_location_service_unsupported_and_timeout(self):
        import unittest.mock as mock

        fallback = mock.Mock()
        expected = mock.sentinel.location
        fallback.locate.return_value = expected
        with mock.patch("call_book.services.location_service.sys.platform", "linux"):
            self.assertIs(LocationService(fallback).locate(), expected)
        with (
            mock.patch("call_book.services.location_service.sys.platform", "win32"),
            mock.patch(
                "call_book.services.location_service.subprocess.run",
                side_effect=__import__("subprocess").TimeoutExpired("powershell", 12),
            ),
        ):
            self.assertIs(LocationService(fallback).locate(), expected)

    def test_ip_location_parses_alternate_coordinate_names_and_validates_ranges(self):
        service = IpLocationService("https://location.example/v1")
        response = mock.MagicMock()
        response.status = 200
        response.headers.get_content_type.return_value = "application/json"
        response.read.return_value = b'{"lat":"44.4268","lng":"26.1025"}'
        response.__enter__.return_value = response
        with (
            mock.patch(
                "call_book.services.location_service.socket.getaddrinfo",
                return_value=[(None, None, None, None, ("127.0.0.1", 443))],
            ),
            mock.patch("call_book.services.location_service.socket.create_connection"),
            mock.patch("call_book.services.location_service.urlopen", return_value=response),
        ):
            result = service.locate()
        self.assertEqual((result.latitude, result.longitude), (44.4268, 26.1025))
        response.read.return_value = b'{"latitude":91,"longitude":0}'
        with (
            mock.patch(
                "call_book.services.location_service.socket.getaddrinfo",
                return_value=[(None, None, None, None, ("127.0.0.1", 443))],
            ),
            mock.patch("call_book.services.location_service.socket.create_connection"),
            mock.patch("call_book.services.location_service.urlopen", return_value=response),
            self.assertRaises(LocationResponseError),
        ):
            service.locate()

    def test_ip_location_reports_dns_failure(self):
        import socket

        with (
            mock.patch("call_book.services.location_service.socket.getaddrinfo", side_effect=socket.gaierror),
            self.assertRaises(LocationDnsError),
        ):
            IpLocationService().locate()


if __name__ == "__main__":
    unittest.main()
